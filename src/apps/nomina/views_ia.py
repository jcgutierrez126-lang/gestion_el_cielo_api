import base64
import json
import os
import logging
import re
import time
import unicodedata
from datetime import date, timedelta
from io import BytesIO

import anthropic
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework import status

from .models import Empleado, TipoLabor, TipoCobro, ControlSemanal
from apps.produccion.models import Lote
from apps.finanzas.models import Egreso, Cuenta

logger = logging.getLogger(__name__)


def _call_openai_vision(system_prompt: str, user_prompt: str, image_b64: str, media_type: str, max_tokens: int = 8192) -> tuple:
    """Llama a GPT-4.1 con visión. Devuelve (texto, tokens_usados)."""
    import openai
    api_key = os.getenv('OPENAI_API_KEY', '')
    client = openai.OpenAI(api_key=api_key, timeout=180.0)
    response = client.chat.completions.create(
        model="o4-mini",
        max_completion_tokens=max_tokens,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{image_b64}"}},
                {"type": "text", "text": user_prompt},
            ]},
        ],
    )
    text = response.choices[0].message.content.strip()
    tokens = (response.usage.prompt_tokens or 0) + (response.usage.completion_tokens or 0)
    return text, tokens


def _claude_create(client, max_retries=3, **kwargs):
    for attempt in range(1, max_retries + 1):
        try:
            return client.messages.create(**kwargs)
        except anthropic.APIStatusError as e:
            if e.status_code == 529 and attempt < max_retries:
                wait = 5 * attempt
                logger.warning('Claude sobrecargado (529), reintento %d/%d en %ds', attempt, max_retries, wait)
                time.sleep(wait)
            else:
                raise


# Correcciones de OCR frecuentes en manuscrito
# SOLO para letras que se confunden visualmente, NO para abreviaturas válidas del modelo
OCR_CORRECCIONES_LABOR = {
    "DS": "DB",   # B se confunde con S en manuscrito
    "DE": "DB",   # variante común
}

# Dígitos que el OCR confunde con letras en abreviaturas manuscritas
_DIGIT_A_LETRA = str.maketrans("0168", "OIGB")


def _corregir_abreviatura(valor):
    """Reemplaza dígitos visualmente similares a letras en abreviaturas de lote/labor."""
    if not valor:
        return valor
    return valor.translate(_DIGIT_A_LETRA)


def _normalizar(texto: str) -> str:
    """Minúsculas, sin acentos, solo letras y espacios."""
    t = unicodedata.normalize('NFD', texto.lower())
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z\s]', '', t).strip()


def _score_similitud(a: str, b: str) -> float:
    """Score 0..1 de similitud entre dos nombres normalizados por tokens."""
    ta = set(_normalizar(a).split())
    tb = set(_normalizar(b).split())
    if not ta or not tb:
        return 0.0
    interseccion = ta & tb
    return len(interseccion) / max(len(ta), len(tb))


def _get_lotes_dict() -> dict:
    """Devuelve {abreviatura: nombre} desde el modelo Lote."""
    try:
        return {
            lote['abreviatura']: lote['nombre']
            for lote in Lote.objects.filter(activo=True).values('abreviatura', 'nombre')
            if lote['abreviatura']
        }
    except Exception:
        return {}


def _get_labores_dict() -> dict:
    """Devuelve {abreviatura: nombre} desde el modelo TipoLabor."""
    try:
        return {
            tl['abreviatura']: tl['nombre']
            for tl in TipoLabor.objects.filter(activo=True).values('abreviatura', 'nombre')
            if tl['abreviatura']
        }
    except Exception:
        return {}


def _get_empleados_activos() -> list:
    """Devuelve lista de nombres completos de empleados activos."""
    try:
        return list(
            Empleado.objects.filter(activo=True)
            .order_by('nombre_completo')
            .values_list('nombre_completo', flat=True)
        )
    except Exception:
        return []


def _get_cobros_dict() -> dict:
    """Devuelve {abreviatura: nombre} desde el modelo TipoCobro."""
    try:
        return {
            tc['abreviatura']: tc['nombre']
            for tc in TipoCobro.objects.all().values('abreviatura', 'nombre')
            if tc['abreviatura']
        }
    except Exception:
        return {"K": "Kilos", "J": "Jornal", "C": "Contrato", "N": "Nómina"}

MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def semana_ref_desde_fecha(fecha_str: str) -> str:
    try:
        d = date.fromisoformat(fecha_str)
    except (ValueError, TypeError):
        return fecha_str or ""
    lunes = d - timedelta(days=d.weekday())
    sabado = lunes + timedelta(days=5)
    mes = MESES_ES[lunes.month - 1]
    if lunes.month == sabado.month:
        return f"Semana del {lunes.day} al {sabado.day} de {mes} de {lunes.year}"
    mes2 = MESES_ES[sabado.month - 1]
    return f"Semana del {lunes.day} de {mes} al {sabado.day} de {mes2} de {lunes.year}"


def _build_system_prompt_diaria() -> str:
    lotes = _get_lotes_dict()
    labores = _get_labores_dict()
    cobros = _get_cobros_dict()
    empleados = _get_empleados_activos()

    lotes_txt = ", ".join(f'"{k}"="{v}"' for k, v in sorted(lotes.items())) if lotes else "M=La Milagrosa, T=El Tanque, LL=El Llano, GU=Guaduas, SJ=San José, N=El Niño"
    labores_txt = ", ".join(f'"{k}"="{v}"' for k, v in sorted(labores.items())) if labores else "R=Recolección, G=Guadaña, A=Abono, B=Banano, E=Embolsada, AT=Aux.Transporte"
    cobros_txt = ", ".join(f'"{k}"="{v}"' for k, v in sorted(cobros.items())) if cobros else "K=Kilos, J=Jornal, C=Contrato, N=Nómina"
    empleados_txt = ", ".join(f'"{e}"' for e in empleados) if empleados else "(sin lista)"

    return f"""Eres un asistente de extracción de datos para la plataforma de gestión de Finca El Cielo.
Tu única tarea es leer imágenes de planillas semanales manuscritas y devolver un JSON con UN REGISTRO POR TRABAJADOR POR DÍA.

REGLAS ESTRICTAS:
- Devuelve SOLO JSON válido, sin texto antes ni después, sin bloques de código.
- Si un campo está en blanco o ilegible, usa null.
- Fechas en formato YYYY-MM-DD.
- Valores numéricos sin separador de miles (ejemplo: 71667 no 71.667, 94000 no 94.000).
- El campo "dia" debe ser exactamente: Lunes, Martes, Miércoles, Jueves, Viernes o Sábado.

CANTIDADES (Col 4, 7, 10, 13, 16, 19 — Cant. de cada día):
- Son siempre dígitos. Lee con extremo cuidado "1" vs "7":
  - "1" = trazo vertical simple o con base → es UNO.
  - "7" = trazo con ángulo oblicuo en la parte superior → es SIETE.
  - Un trazo casi vertical con solo una pequeña inclinación es 1, NO 7.
  - En esta planilla, los trabajadores NO suelen anotar cantidades de 7 unidades cuando hacen una sola labor; 1 es mucho más frecuente que 7.
  - Si dudas entre 1 y 7, prefiere 1.

LOTES — copia la ABREVIATURA EXACTA como está escrita. NO conviertas a nombre completo. Si en blanco → null.
Abreviaturas del modelo: {lotes_txt}
Ejemplos reales: "GD", "LL", "B", "F", "BLB", "LLB", "SF", "Sh", "Nn", "ML", "BL", "Es", "SJ", "G"
IMPORTANTE: Las abreviaturas de lotes y labores contienen SOLO LETRAS, nunca dígitos.
Si ves un dígito es un error de lectura — reemplázalo por la letra visualmente similar: 6→G, 0→O, 1→I.

LABORES — copia la ABREVIATURA EXACTA como está escrita. NO conviertas a nombre completo.
Abreviaturas del modelo: {labores_txt}
Ejemplos reales: "Rc", "PL", "Gn", "FR", "AR", "Yr", "AT", "EB", "PJ", "DJ", "MH", "AL", "DM", "VR", "DB"
IMPORTANTE: "AR"=Arriero (labor válida), "VR"=Varios — son abreviaturas DISTINTAS. No corrijas "AR" a "VR".
Confusiones frecuentes en manuscrito:
- B se confunde con S/E → "DS"→"DB", "DE"→"DB"
- D se confunde con A → "AB"→"DB" (la letra D manuscrita parece A con el arco abierto)
- V se confunde con Y → si ves "Yr", considera si la abreviatura real es "Vr" (VR=Varios). Lee el contexto.
- D se confunde con P → "DJ" y "PJ" son labores distintas.
- A se confunde con F → "FR"→"AR" (Arriado): si ves "FR" y la fila no es claramente Fruta, considera "AR".
  Clave: "AR"=Arriero (movimiento de mulas/carga), "FR"=Fruta (cosecha de fruta). Son actividades distintas. Lee el contexto del lote.
- Los dígitos son errores OCR de letras: 6→G, 0→O, 1→I, 8→B.

TIPO DE COBRO — columna K·J / C·N (col 21) — lee con extremo cuidado:
- La letra "J" (palo vertical + gancho) = Jornal → tipo_cobro="jornal"
- La letra "N" (dos palos y diagonal) = Nómina → tipo_cobro="nomina"
- Son MUY distintas. "J" tiene forma de anzuelo, "N" tiene dos trazos verticales.
- La letra "K" = Kilos → tipo_cobro="kilos"
- La letra "C" = Contrato → tipo_cobro="contrato"
- NUNCA pongas "nomina" si la letra en la planilla se parece a "J".

TIPOS DE COBRO (letra al final de la fila): {cobros_txt}
IMPORTANTE: "N"=Nómina ≠ "J"=Jornal. Son letras distintas. Lee con cuidado la columna K·J/C·N.
Para tipo_cobro usa en minúsculas: kilos, jornal, contrato, nomina.

TRABAJADORES ACTIVOS DE LA FINCA (lista oficial):
{empleados_txt}

Para el campo "nombre":
- Busca el nombre de la planilla en la lista oficial de arriba.
- Si hay una coincidencia clara (mismo apellido, mismo primer nombre) → usa EL NOMBRE EXACTO de la lista oficial.
- Si el nombre está ilegible o truncado → usa el de la lista que más se parezca fonéticamente.
- NUNCA inventes nombres que no estén en la lista. Si no hay match razonable → transcribe lo que ves literalmente.
- Los números en los nombres son errores de lectura: "R1goberto" → "Rigoberto".

REGLA ABSOLUTA DE AISLAMIENTO DE FILAS:
- Cada línea horizontal de la tabla = UN trabajador y SOLO ese trabajador.
- Los datos de una fila JAMÁS se mezclan con la fila de arriba ni con la de abajo.
- El nombre al inicio de la fila es el "propietario" de TODOS los datos de esa misma línea horizontal.
- Si una celda parece pertenecer a otra fila, es un error de lectura — asígnala al trabajador de esa fila.
- El COBRO (col 21) y el VALOR (col 22) al final pertenecen EXCLUSIVAMENTE a esa fila.
- Omite los días donde el trabajador no tiene labor registrada.
- Gastos/compras van en "observaciones", NO como registros de trabajador.
"""


USER_PROMPT_DIARIA = """Extrae todos los datos de esta planilla SEMANAL de trabajadores de Finca El Cielo.

FORMULARIO IMPRESO — ESTRUCTURA FIJA (la foto es siempre de este mismo formulario):
22 columnas en este orden exacto:

  Col  1 : NOMBRE del trabajador
  Col  2 : Labor  Lunes      ← primera sub-col del grupo Lunes (separada por línea gruesa)
  Col  3 : Lote   Lunes
  Col  4 : Cant.  Lunes
  Col  5 : Labor  Martes     ← primera sub-col del grupo Martes (línea gruesa)
  Col  6 : Lote   Martes
  Col  7 : Cant.  Martes
  Col  8 : Labor  Miércoles  ← línea gruesa
  Col  9 : Lote   Miércoles
  Col 10 : Cant.  Miércoles
  Col 11 : Labor  Jueves     ← línea gruesa
  Col 12 : Lote   Jueves
  Col 13 : Cant.  Jueves
  Col 14 : Labor  Viernes    ← línea gruesa
  Col 15 : Lote   Viernes
  Col 16 : Cant.  Viernes
  Col 17 : Labor  Sábado     ← línea gruesa
  Col 18 : Lote   Sábado
  Col 19 : Cant.  Sábado
  Col 20 : 1/2    Sábado     (medio día — ignora, no extraer)
  Col 21 : COBRO             (K / J / C / N — aplica a toda la semana de esa fila)
  Col 22 : VALOR             (cifra — aplica a toda la semana de esa fila)

REGLA FUNDAMENTAL: si la celda Labor (col 2,5,8,11,14,17) de un día está VACÍA → ese trabajador
NO trabajó ese día → NO crees registro para ese día. Solo crea registro cuando Labor tiene algo escrito.
El COBRO y VALOR al final son de ESA FILA únicamente, no de la fila de arriba ni abajo.
No hay domingo. Máximo 6 registros por trabajador (Lunes→Sábado).

INSTRUCCIONES:
1. Crea UN REGISTRO por trabajador-día que tenga labor registrada.
2. El encabezado tiene "Del DD/MM al DD/MM de YYYY" — el primer DD/MM es el LUNES = fecha_inicio (YYYY-MM-DD).
   Fecha de cada día: lunes=fecha_inicio, martes=+1d, miércoles=+2d, jueves=+3d, viernes=+4d, sábado=+5d.
3. La CANTIDAD es Col 4/7/10/13/16/19 (tercera de cada grupo). Si en blanco → null.
4. El COBRO es Col 21. Léela con cuidado:
   - K → tipo_cobro="kilos"    (N es Nómina, NO Jornal)
   - J → tipo_cobro="jornal"
   - C → tipo_cobro="contrato"
   - N → tipo_cobro="nomina"

5. CÓMO ASIGNAR EL VALOR de cada registro diario:

   La columna VALOR al final de la fila es siempre una TARIFA o PRECIO que se repite igual para cada día.
   NUNCA dividas el valor entre días trabajados.

   Cobro K (kilos):
   - VALOR fila = PRECIO POR UNIDAD (ej: 1.300 = $1.300/unidad, 1.450 = $1.450/unidad).
   - Para cada día: en el campo "valor" del JSON pon EL PRECIO UNITARIO (lo que está escrito en la planilla).
     NO calcules cantidad × precio. El sistema lo calcula solo.
   - Si cantidad en blanco → cantidad = 1 (un solo bulto/unidad).
   - Ejemplo correcto: cantidad=106, valor=1300 (precio por unidad, NO 137800).

   Cobro J (jornal), N (nómina) o C (contrato):
   - VALOR fila = TARIFA DIARIA del trabajador (ej: 74.000 = $74.000/día).
   - Para CADA día que trabajó: valor = esa misma tarifa diaria.
   - Ejemplo: trabajó 3 días con tarifa 74.000 → 3 registros, cada uno con valor=74000.
   - NUNCA dejes valor null.

6. Quita los puntos de miles en todos los números: 71.667→71667, 1.300→1300, 60.000→60000.
7. Para el lote: la abreviatura de la columna LOTE de ESE DÍA. Si en blanco → null.
8. En "observaciones" incluye gastos, vales y notas al pie.

Devuelve exactamente este JSON:
{
  "fecha_inicio": "YYYY-MM-DD del lunes de la semana",
  "semana_ref": "texto del encabezado tal como aparece",
  "valor_jornal": número_o_null,
  "registros": [
    {
      "nombre": "nombre tal como está escrito en la planilla",
      "dia": "Lunes|Martes|Miércoles|Jueves|Viernes|Sábado",
      "fecha": "YYYY-MM-DD de ese día",
      "lote": "abreviatura del lote (ej: GD, LL, B, SF, Sh, BL, ML, Es, Nn, SJ, G) — SOLO LETRAS, nunca dígitos. o null",
      "labor": "abreviatura de la labor tal como está escrita en la planilla (ej: Rc, PL, Gn, FR, EB, DM, AT, AR, MH, AL, VR, DB) o null",
      "cantidad": número_o_null,
      "tipo_cobro": "jornal|kilos|contrato|nomina",
      "valor": número_o_null
    }
  ],
  "observaciones": "texto o null"
}
"""


# ── Prompt semanal (formato antiguo, compatibilidad) ──────────────────────────

def _build_system_prompt_semanal() -> str:
    lotes = _get_lotes_dict()
    labores = _get_labores_dict()
    empleados = _get_empleados_activos()
    lotes_txt = ", ".join(f'"{k}"="{v}"' for k, v in sorted(lotes.items())) if lotes else ""
    labores_txt = ", ".join(f'"{k}"="{v}"' for k, v in sorted(labores.items())) if labores else ""
    empleados_txt = ", ".join(f'"{e}"' for e in empleados) if empleados else ""
    return f"""Eres un asistente de extracción de datos para Finca El Cielo.
Devuelve SOLO JSON válido. Si un campo está en blanco usa null.
LOTES: {lotes_txt}
LABORES: {labores_txt}
TRABAJADORES: {empleados_txt}
Para tipo_cobro usa: kilos, jornal, contrato, nomina. Fechas YYYY-MM-DD. Números sin separadores de miles.
"""

USER_PROMPT = """Extrae todos los datos de esta planilla semanal de Finca El Cielo.
Devuelve este JSON exacto:
{
  "semana": {"fecha_inicio": "YYYY-MM-DD o null", "fecha_fin": "YYYY-MM-DD o null"},
  "registros": [
    {
      "trabajador": "nombre tal como aparece",
      "cedula": "número o null",
      "lote": "nombre del lote o null",
      "tipo_labor": "nombre de la labor",
      "tipo_cobro": "kilos|jornal|contrato|nomina",
      "kilos": número_o_null,
      "jornales": número_o_null,
      "costo_unidad": número_o_null,
      "valor": número_o_null
    }
  ],
  "observaciones": "texto o null"
}
"""


class LeerPlanillaDiariaView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        imagen = request.FILES.get('imagen')
        if not imagen:
            return Response({'error': 'Se requiere el campo "imagen".'}, status=status.HTTP_400_BAD_REQUEST)

        proveedor = request.data.get('proveedor_ia', 'claude')

        contenido = imagen.read()
        media_type = imagen.content_type or 'image/jpeg'
        SOPORTADOS = ('image/jpeg', 'image/png', 'image/webp', 'image/gif')
        if media_type not in SOPORTADOS:
            if media_type in ('image/heic', 'image/heif'):
                try:
                    import pillow_heif
                    from PIL import Image as PILImage
                    pillow_heif.register_heif_opener()
                    img = PILImage.open(BytesIO(contenido))
                    buf = BytesIO()
                    img.save(buf, format='JPEG', quality=90)
                    contenido = buf.getvalue()
                    media_type = 'image/jpeg'
                except Exception:
                    return Response(
                        {'error': 'Formato de imagen no soportado (HEIC). Toma la foto en JPG o PNG.'},
                        status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                    )
            else:
                media_type = 'image/jpeg'

        imagen_b64 = base64.standard_b64encode(contenido).decode('utf-8')
        raw = ''

        try:
            if proveedor == 'gpt':
                raw, tokens_usados = _call_openai_vision(
                    _build_system_prompt_diaria(), USER_PROMPT_DIARIA,
                    imagen_b64, media_type, max_tokens=8192,
                )
            else:
                api_key = os.getenv('ANTHROPIC_API_KEY', '')
                if not api_key:
                    return Response({'error': 'ANTHROPIC_API_KEY no configurada.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                client = anthropic.Anthropic(api_key=api_key)
                message = _claude_create(
                    client,
                    model='claude-opus-4-7',
                    max_tokens=8192,
                    system=_build_system_prompt_diaria(),
                    messages=[{'role': 'user', 'content': [
                        {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': imagen_b64}},
                        {'type': 'text', 'text': USER_PROMPT_DIARIA},
                    ]}],
                )
                if not message.content:
                    return Response({'error': 'El modelo no generó respuesta.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
                raw = message.content[0].text.strip()
                tokens_usados = message.usage.input_tokens + message.usage.output_tokens

            logger.debug('IA raw response (primeros 500): %s', raw[:500])

            if not raw:
                return Response({'error': 'El modelo devolvió una respuesta vacía.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1]
                raw = raw.rsplit('```', 1)[0]

            datos = json.loads(raw)

            for reg in datos.get('registros', []):
                reg['lote'] = _corregir_abreviatura(reg.get('lote'))
                reg['labor'] = _corregir_abreviatura(reg.get('labor'))

            fecha = datos.get('fecha_inicio') or datos.get('fecha') or ''
            if not datos.get('semana_ref'):
                datos['semana_ref'] = semana_ref_desde_fecha(fecha)
            if not datos.get('fecha') and fecha:
                datos['fecha'] = fecha

            return Response({'ok': True, 'datos': datos, 'tokens_usados': tokens_usados})

        except json.JSONDecodeError as e:
            logger.error('IA devolvió JSON inválido: %s', raw[:500])
            return Response({'error': 'No se pudo parsear la respuesta de la IA.', 'detalle': str(e)},
                            status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        except anthropic.APIError as e:
            logger.error('Error API Anthropic: %s', e)
            return Response({'error': 'Error al llamar a la API de Claude.', 'detalle': str(e)},
                            status=status.HTTP_502_BAD_GATEWAY)
        except Exception as e:
            logger.error('Error IA: %s', e)
            return Response({'error': f'Error del servicio de IA: {e}'}, status=status.HTTP_502_BAD_GATEWAY)


class LeerPlanillaSemanalExcelView(APIView):
    """Parsea el Excel de planilla semanal (tab Labores) y devuelve registros para revisión."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    # Columnas 0-indexed por día: (lote, labor, cob, cant)
    _DIA_COLS = [
        (1, 2, 3, 4),      # Lunes:     B C D E
        (5, 6, 7, 8),      # Martes:    F G H I
        (9, 10, 11, 12),   # Miércoles: J K L M
        (13, 14, 15, 16),  # Jueves:    N O P Q
        (17, 18, 19, 20),  # Viernes:   R S T U
        (21, 22, 23, 24),  # Sábado:    V W X Y
    ]
    _DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    _COL_MEDIO_DIA = 25   # Z  — "X" = sábado medio día
    _COL_PRECIO    = 27   # AB — $ Jornal / $ Kilo
    _COL_TIPO      = 28   # AC — K=kilos, C=contrato, vacío=jornal

    @staticmethod
    def _cel(row, idx):
        """Extrae celda como string limpio, devuelve '' si vacía."""
        if idx < len(row) and row[idx] is not None:
            v = str(row[idx]).strip()
            return v if v.lower() not in ('none', 'nan') else ''
        return ''

    @staticmethod
    def _num(s):
        """Convierte string a float; devuelve None si no parseable."""
        try:
            return float(s.replace(',', '.')) if s else None
        except (ValueError, AttributeError):
            return None

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'Se requiere el campo "archivo".'}, status=status.HTTP_400_BAD_REQUEST)

        fecha_inicio_str = request.data.get('fecha_inicio', '')
        try:
            fecha_lunes = date.fromisoformat(fecha_inicio_str)
        except (ValueError, TypeError):
            hoy = date.today()
            fecha_lunes = hoy - timedelta(days=hoy.weekday())

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(archivo.read()), data_only=True)
            ws = wb['Labores']
        except Exception as e:
            return Response({'error': f'No se pudo leer el archivo: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        registros = []

        for row in ws.iter_rows(min_row=6, max_row=25, values_only=True):
            nombre = self._cel(row, 0)
            if not nombre:
                continue

            tipo_raw = self._cel(row, self._COL_TIPO).upper()
            if tipo_raw == 'K':
                tipo_cobro = 'kilos'
            elif tipo_raw == 'C':
                tipo_cobro = 'contrato'
            else:
                tipo_cobro = 'jornal'

            precio = self._num(self._cel(row, self._COL_PRECIO)) or 0
            medio_dia_sabado = self._cel(row, self._COL_MEDIO_DIA).upper() == 'X'

            for i, (ci_lote, ci_labor, _ci_cob, ci_cant) in enumerate(self._DIA_COLS):
                lote_val  = self._cel(row, ci_lote)
                labor_val = self._cel(row, ci_labor)
                cant_str  = self._cel(row, ci_cant)

                if not lote_val and not labor_val and not cant_str:
                    continue

                es_medio_dia = (i == 5) and medio_dia_sabado
                cant_num     = self._num(cant_str)
                fecha_dia    = fecha_lunes + timedelta(days=i)

                if tipo_cobro == 'kilos':
                    cantidad = cant_num
                    valor    = round(cant_num * precio) if cant_num and precio else None
                elif tipo_cobro == 'jornal':
                    cantidad = 0.5 if es_medio_dia else 1.0
                    valor    = round(precio * cantidad) if precio else None
                else:  # contrato — valor se calcula globalmente, aquí queda None
                    cantidad = cant_num
                    valor    = None

                registros.append({
                    'nombre':     nombre,
                    'dia':        self._DIAS[i],
                    'fecha':      fecha_dia.isoformat(),
                    'lote':       lote_val or None,
                    'labor':      labor_val or None,
                    'cantidad':   cantidad,
                    'tipo_cobro': tipo_cobro,
                    'valor':      valor,
                })

        semana_ref = semana_ref_desde_fecha(fecha_lunes.isoformat())

        return Response({
            'ok': True,
            'datos': {
                'fecha_inicio': fecha_lunes.isoformat(),
                'semana_ref':   semana_ref,
                'registros':    registros,
                'observaciones': None,
            },
        })


class LeerPlanillaView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        imagen = request.FILES.get('imagen')
        if not imagen:
            return Response({'error': 'Se requiere el campo "imagen".'}, status=status.HTTP_400_BAD_REQUEST)

        proveedor = request.data.get('proveedor_ia', 'claude')

        contenido = imagen.read()
        media_type = imagen.content_type or 'image/jpeg'
        SOPORTADOS = ('image/jpeg', 'image/png', 'image/webp', 'image/gif')
        if media_type not in SOPORTADOS:
            if media_type in ('image/heic', 'image/heif'):
                try:
                    import pillow_heif
                    from PIL import Image as PILImage
                    pillow_heif.register_heif_opener()
                    img = PILImage.open(BytesIO(contenido))
                    buf = BytesIO()
                    img.save(buf, format='JPEG', quality=90)
                    contenido = buf.getvalue()
                    media_type = 'image/jpeg'
                except Exception:
                    return Response(
                        {'error': 'Formato de imagen no soportado (HEIC). Toma la foto en JPG o PNG.'},
                        status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                    )
            else:
                media_type = 'image/jpeg'

        imagen_b64 = base64.standard_b64encode(contenido).decode('utf-8')
        raw = ''

        try:
            if proveedor == 'gpt':
                raw, tokens_usados = _call_openai_vision(
                    _build_system_prompt_semanal(), USER_PROMPT,
                    imagen_b64, media_type, max_tokens=4096,
                )
            else:
                api_key = os.getenv('ANTHROPIC_API_KEY', '')
                if not api_key:
                    return Response({'error': 'ANTHROPIC_API_KEY no configurada.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                client = anthropic.Anthropic(api_key=api_key)
                message = _claude_create(
                    client,
                    model='claude-opus-4-7',
                    max_tokens=4096,
                    system=_build_system_prompt_semanal(),
                    messages=[{'role': 'user', 'content': [
                        {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': imagen_b64}},
                        {'type': 'text', 'text': USER_PROMPT},
                    ]}],
                )
                if not message.content:
                    return Response({'error': 'El modelo no generó respuesta.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
                raw = message.content[0].text.strip()
                tokens_usados = message.usage.input_tokens + message.usage.output_tokens

            if not raw:
                return Response({'error': 'El modelo devolvió una respuesta vacía.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1]
                raw = raw.rsplit('```', 1)[0]

            datos = json.loads(raw)
            return Response({'ok': True, 'datos': datos, 'tokens_usados': tokens_usados})

        except json.JSONDecodeError as e:
            logger.error('IA devolvió JSON inválido: %s', raw[:500])
            return Response({'error': 'No se pudo parsear la respuesta de la IA.', 'detalle': str(e)},
                            status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        except anthropic.APIError as e:
            logger.error('Error API Anthropic: %s', e)
            return Response({'error': 'Error al llamar a la API de Claude.', 'detalle': str(e)},
                            status=status.HTTP_502_BAD_GATEWAY)
        except Exception as e:
            logger.error('Error IA: %s', e)
            return Response({'error': f'Error del servicio de IA: {e}'}, status=status.HTTP_502_BAD_GATEWAY)


def _buscar_empleado(nombre: str, umbral: float = 0.4):
    """Busca el empleado más similar usando score de tokens. Umbral 0..1."""
    if not nombre:
        return None
    nombre = nombre.strip()
    # Match exacto primero
    try:
        return Empleado.objects.get(nombre_completo__iexact=nombre)
    except (Empleado.DoesNotExist, Empleado.MultipleObjectsReturned):
        pass
    # Match por fragmento
    qs = Empleado.objects.filter(nombre_completo__icontains=nombre.split()[0] if nombre.split() else nombre)
    if qs.count() == 1:
        return qs.first()
    # Fuzzy por score de tokens sobre todos los activos
    todos = list(Empleado.objects.filter(activo=True).values('id', 'nombre_completo'))
    mejor, mejor_score = None, 0.0
    for e in todos:
        score = _score_similitud(nombre, e['nombre_completo'])
        if score > mejor_score:
            mejor_score = score
            mejor = e
    if mejor_score >= umbral:
        return Empleado.objects.get(id=mejor['id'])
    return None


def _buscar_tipo_labor(texto: str):
    if not texto:
        return None
    t = texto.strip()
    t_corr = OCR_CORRECCIONES_LABOR.get(t.upper(), t)
    return (
        TipoLabor.objects.filter(abreviatura__iexact=t_corr).first() or
        TipoLabor.objects.filter(abreviatura__iexact=t).first() or
        TipoLabor.objects.filter(nombre__icontains=t).first()
    )


def _buscar_tipo_cobro(texto: str):
    if not texto:
        return None
    t = texto.strip()
    return (
        TipoCobro.objects.filter(abreviatura__iexact=t).first() or
        TipoCobro.objects.filter(nombre__icontains=t).first()
    )


def _buscar_lote(texto: str):
    if not texto:
        return None
    t = texto.strip()
    return (
        Lote.objects.filter(abreviatura__iexact=t).first() or
        Lote.objects.filter(nombre__icontains=t).first()
    )


class GuardarPlanillaView(APIView):
    """Guarda registros parseados de planilla en ControlSemanal resolviendo FKs por nombre/abreviatura."""
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser]

    def post(self, request):
        datos = request.data
        semana_ref = datos.get('semana_ref', '')
        fecha_inicio_str = datos.get('fecha_inicio', '')
        registros_raw = datos.get('registros', [])
        valor_jornal_global = datos.get('valor_jornal')

        try:
            fecha_lunes = date.fromisoformat(fecha_inicio_str)
        except (ValueError, TypeError):
            return Response({'error': 'fecha_inicio inválida (YYYY-MM-DD requerido)'}, status=status.HTTP_400_BAD_REQUEST)

        fecha_sabado = fecha_lunes + timedelta(days=5)
        creados = []
        errores = []

        for i, r in enumerate(registros_raw):
            nombre = r.get('nombre', '')
            labor_txt = r.get('labor', '')
            tipo_cobro_txt = r.get('tipo_cobro', '')
            lote_txt = r.get('lote') or ''
            dia = r.get('dia', '')
            fecha_str = r.get('fecha', '') or fecha_inicio_str
            cantidad = r.get('cantidad')
            valor = r.get('valor')
            costo_unidad = r.get('valor_jornal') or valor_jornal_global or 0

            empleado = _buscar_empleado(nombre)
            tipo_labor = _buscar_tipo_labor(labor_txt)
            tipo_cobro = _buscar_tipo_cobro(tipo_cobro_txt)
            lote = _buscar_lote(lote_txt)

            faltantes = []
            if not empleado:
                faltantes.append(f'empleado "{nombre}"')
            if not tipo_labor:
                faltantes.append(f'labor "{labor_txt}"')
            if not tipo_cobro:
                faltantes.append(f'tipo_cobro "{tipo_cobro_txt}"')

            if faltantes:
                errores.append({'fila': i + 1, 'nombre': nombre, 'motivo': f'No encontrado: {", ".join(faltantes)}'})
                continue

            try:
                fecha_dia = date.fromisoformat(fecha_str)
            except (ValueError, TypeError):
                fecha_dia = fecha_lunes

            kilos = float(cantidad) if tipo_cobro_txt.lower() == 'kilos' and cantidad is not None else None
            jornales = float(cantidad) if tipo_cobro_txt.lower() != 'kilos' and cantidad is not None else None

            ControlSemanal.objects.create(
                empleado=empleado,
                semana_ref=semana_ref,
                dia=dia,
                fecha=fecha_dia,
                fecha_inicio=fecha_lunes,
                fecha_fin=fecha_sabado,
                tipo_labor=tipo_labor,
                tipo_cobro=tipo_cobro,
                lote=lote,
                kilos=kilos,
                jornales=jornales,
                costo_unidad=costo_unidad,
                valor=valor or 0,
            )
            creados.append(nombre)

        # Crear egreso de nómina automáticamente si hubo registros guardados
        egreso_creado = False
        if creados:
            total_nomina = sum(
                ControlSemanal.objects.filter(
                    semana_ref=semana_ref,
                    fecha_inicio=fecha_lunes,
                ).values_list('valor', flat=True)
            )
            cuenta = (
                Cuenta.objects.filter(tipo='agencia').first()
                or Cuenta.objects.first()
            )
            if cuenta and total_nomina:
                Egreso.objects.update_or_create(
                    nombre=f'Nómina {semana_ref}',
                    fecha=fecha_lunes,
                    defaults={
                        'cuenta': cuenta,
                        'categoria': 'nomina',
                        'valor': total_nomina,
                        'descripcion': f'{len(creados)} registros — {semana_ref}',
                        'estado': 'pagada',
                    }
                )
                egreso_creado = True

        return Response({
            'ok': True,
            'creados': len(creados),
            'errores': errores,
            'semana_ref': semana_ref,
            'egreso_creado': egreso_creado,
        }, status=status.HTTP_201_CREATED)


class MatchEmpleadoView(APIView):
    """Recibe un nombre extraído por OCR y devuelve los empleados más similares del modelo."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        nombre = (request.data.get('nombre') or '').strip()
        if not nombre:
            return Response({'error': 'Se requiere "nombre".'}, status=status.HTTP_400_BAD_REQUEST)

        todos = list(
            Empleado.objects.filter(activo=True)
            .order_by('nombre_completo')
            .values('id', 'nombre_completo')
        )
        scored = [
            {'id': e['id'], 'nombre': e['nombre_completo'], 'score': _score_similitud(nombre, e['nombre_completo'])}
            for e in todos
        ]
        scored.sort(key=lambda x: x['score'], reverse=True)
        return Response({'candidatos': scored[:10]})


class ListEmpleadosActivosView(APIView):
    """Lista todos los empleados activos para el frontend (match manual)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        empleados = list(
            Empleado.objects.filter(activo=True)
            .order_by('nombre_completo')
            .values('id', 'nombre_completo')
        )
        return Response({'empleados': empleados})
