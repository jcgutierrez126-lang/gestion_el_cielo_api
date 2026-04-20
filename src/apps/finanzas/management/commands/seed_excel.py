"""
Management command: seed_excel

Usage:
    python manage.py seed_excel /tmp/finca_la_holanda.xlsx

Populates the database from the historical Excel spreadsheet.
Skips Control Semanal. Uses read_only=True to minimize RAM usage.
"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.finanzas.models import (
    Cuenta, Egreso, Ingreso, Observacion, Proveedor, Transaccion,
)
from apps.nomina.models import AbonoPrestamo, Empleado, PrestamoEmpleado
from apps.produccion.models import (
    Floracion, Lote, MezclaAbono, MezclaAbonoFertilizante,
    VentaBanano, VentaCafe, VentaCafeTostado,
)

# ─────────────────────────── helpers ────────────────────────────────────────

def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _dec(v, default=None) -> Decimal | None:
    if v is None:
        return default
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return default


def _date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _norm(s: str) -> str:
    """Lower-case, strip accent lookalikes, collapse spaces."""
    return (
        s.lower().strip()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )


# ─────────────────────── choice mappers ─────────────────────────────────────

CATEGORIA_MAP = {
    "viaticos": "viaticos", "viaticos": "viaticos",
    "varios": "varios",
    "fertilizantes": "fertilizantes",
    "herbicidas": "herbicidas",
    "nomina": "nomina",
    "seguridad social": "seguridad_social",
    "transporte": "transporte",
    "acueducto": "acueducto",
    "epm": "epm",
    "comsab": "comsab",
    "mantenimientos": "mantenimientos",
    "beneficio": "beneficio",
    "guadana": "guadana", "guadaña": "guadana",
    "construcciones": "construcciones",
    "impuestos": "impuestos",
    "animales": "animales",
    "siembra": "siembra",
    "herramientas": "herramientas",
    "broca": "broca",
    "roya": "roya",
    "moto": "moto",
    "prestamo empleados": "prestamo_empleados",
    "no aplica": "no_aplica",
    "activos fijos": "activos_fijos",
    "banano": "banano",
    "compra finca": "compra_finca",
    "capacitaciones": "capacitaciones",
    "venta cafe": "venta_cafe",
}

TIPO_CAFE_MAP = {
    "pergamino seco": "pergamino_seco",
    "pasilla": "pasilla",
    "cereza": "cereza",
    "verde": "verde",
    "corriente": "corriente",
}

CALIDAD_MAP = {
    "buena": "buena",
    "regular": "regular",
    "muy buena": "muy_buena",
    "excelente": "excelente",
}


def _map_categoria(raw: str) -> str:
    return CATEGORIA_MAP.get(_norm(raw), "varios")


def _map_tipo_cafe(raw: str) -> str:
    return TIPO_CAFE_MAP.get(_norm(raw), "pergamino_seco")


def _map_calidad(raw: str) -> str:
    return CALIDAD_MAP.get(_norm(raw), "buena")


def _map_tipo_platano(raw: str) -> str:
    if not raw:
        return "banano_extra"
    r = _norm(raw)
    if "africa" in r:
        if "primera" in r:
            return "africa_primera"
        if "segunda" in r:
            return "africa_segunda"
        return "africa_extra"
    if "dominico" in r:
        if "primera" in r:
            return "dominico_primera"
        if "segunda" in r:
            return "dominico_segunda"
        return "dominico_extra"
    if "guineo" in r:
        return "guineo"
    if "harton" in r:
        if "primera" in r:
            return "harton_primera"
        if "segunda" in r:
            return "harton_segunda"
        return "harton_extra"
    if "murrapo" in r:
        if "segunda" in r:
            return "murrapo_segunda"
        return "murrapo_primera"
    if "platano" in r or "plátano" in r:
        if "segunda" in r or "seg" in r:
            return "platano_segunda"
        return "platano_extra"
    if "primera" in r or "especial" in r:
        return "banano_primera"
    if "segunda" in r:
        return "banano_segunda"
    return "banano_extra"


# ─────────────────────── cuenta helper ──────────────────────────────────────

_CUENTA_TIPOS = {
    "bancolombia": "bancaria", "banco": "bancaria",
    "daviplata": "bancaria", "nequi": "bancaria",
    "prestamo": "prestamo", "préstamo": "prestamo",
    "agencia": "agencia", "cooperativa": "agencia",
    "dividendo": "dividendos",
    "vale": "vale",
}


def _cuenta_tipo(nombre: str) -> str:
    n = _norm(nombre)
    for key, tipo in _CUENTA_TIPOS.items():
        if key in n:
            return tipo
    return "efectivo"


_cuenta_cache: dict[str, Cuenta] = {}


def _get_cuenta(nombre: str | None) -> Cuenta | None:
    if not nombre:
        return None
    nombre = nombre.strip()
    if nombre in _cuenta_cache:
        return _cuenta_cache[nombre]
    obj, _ = Cuenta.objects.get_or_create(
        nombre=nombre,
        defaults={"tipo": _cuenta_tipo(nombre), "saldo_inicial": 0},
    )
    _cuenta_cache[nombre] = obj
    return obj


# ─────────────────────── main command ───────────────────────────────────────

REAL_LOTE_NAMES = {
    "La Milagrosa", "La Cruz", "San José", "El Niño", "Santo Domingo",
    "San Charbel", "San Charbel Nuevo", "La Ceja Palos", "La Ceja Produccion",
    "Huerta", "Hoyo Caliente 1", "Hoyo Caliente 2", "Guaduas", "La Estrella",
    "La Bola", "Abejorral", "El Llano", "San Francisco",
    "Banano Holanda", "Banano Milagrosa", "Banano Inmaculada",
}


class Command(BaseCommand):
    help = "Seed database from Finca LA HOLANDA.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to the .xlsx file")
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete existing data before seeding (dangerous!)",
        )

    def handle(self, *args, **options):
        path = options["excel_path"]
        if not os.path.exists(path):
            raise CommandError(f"File not found: {path}")

        self.stdout.write(f"Loading workbook: {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        if options.get("clear"):
            self._clear_data()

        with transaction.atomic():
            self._seed_cuentas_base()
            self._seed_proveedores(wb)
            self._seed_empleados(wb)
            self._seed_lotes(wb)
            self._seed_observaciones(wb)
            self._seed_egresos(wb)
            self._seed_ingresos(wb)
            self._seed_transacciones(wb)
            self._seed_prestamos(wb)
            self._seed_ventas_cafe(wb)
            self._seed_ventas_banano(wb)
            self._seed_floraciones(wb)
            self._seed_mezclas_abono(wb)

        self.stdout.write(self.style.SUCCESS("✓ Seed completado exitosamente."))

    # ── clear ────────────────────────────────────────────────────────────────

    def _clear_data(self):
        self.stdout.write(self.style.WARNING("Borrando datos existentes..."))
        MezclaAbonoFertilizante.objects.all().delete()
        MezclaAbono.objects.all().delete()
        Floracion.objects.all().delete()
        VentaBanano.objects.all().delete()
        VentaCafe.objects.all().delete()
        VentaCafeTostado.objects.all().delete()
        Lote.objects.all().delete()
        AbonoPrestamo.objects.all().delete()
        PrestamoEmpleado.objects.all().delete()
        Empleado.objects.all().delete()
        Observacion.objects.all().delete()
        Egreso.objects.all().delete()
        Ingreso.objects.all().delete()
        Transaccion.objects.all().delete()
        Proveedor.objects.all().delete()
        Cuenta.objects.all().delete()

    # ── cuentas base ─────────────────────────────────────────────────────────

    def _seed_cuentas_base(self):
        base = [
            ("Efectivo", "efectivo"),
            ("Bancolombia Natalia", "bancaria"),
            ("Agencia", "agencia"),
            ("Préstamo Enrique", "prestamo"),
            ("Préstamo Juan Camilo", "prestamo"),
            ("Préstamo Miriam", "prestamo"),
            ("Dividendos", "dividendos"),
            ("Vale", "vale"),
            ("No aplica", "efectivo"),
        ]
        for nombre, tipo in base:
            obj, created = Cuenta.objects.get_or_create(
                nombre=nombre,
                defaults={"tipo": tipo, "saldo_inicial": 0},
            )
            _cuenta_cache[nombre] = obj
        self.stdout.write(f"  Cuentas base: {len(base)}")

    # ── proveedores ───────────────────────────────────────────────────────────

    def _seed_proveedores(self, wb):
        ws = wb["Proveedores"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        created = 0
        for row in rows:
            if not row[0]:
                continue
            nombre = _s(row[0])
            if not nombre:
                continue
            telefono = _s(row[1]) if row[1] else None
            celular = _s(row[2]) if row[2] else None
            cedula = _s(row[3]) if row[3] else None
            direccion = _s(row[4]) if row[4] else None
            ciudad = _s(row[5]) if row[5] else None
            email = _s(row[6]) if row[6] else None
            comentarios = _s(row[7]) if row[7] else None
            Proveedor.objects.get_or_create(
                nombre=nombre,
                defaults=dict(
                    telefono=telefono, celular=celular, cedula_nit=cedula,
                    direccion=direccion, ciudad=ciudad, email=email,
                    comentarios=comentarios,
                ),
            )
            created += 1
        self.stdout.write(f"  Proveedores: {created} procesados")

    # ── empleados ─────────────────────────────────────────────────────────────

    def _seed_empleados(self, wb):
        ws = wb["Empleados"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        created = 0
        for row in rows:
            if not row[0]:
                continue
            nombre = _s(row[0])
            cedula = str(int(row[2])) if row[2] and isinstance(row[2], (int, float)) else (_s(row[2]) or None)
            Empleado.objects.get_or_create(
                nombre_completo=nombre,
                defaults=dict(
                    cedula=cedula,
                    telefono=_s(row[1]) or None,
                    labor=_s(row[3]) or None,
                    jornal=_dec(row[4]),
                    fecha_ingreso=_date(row[5]),
                    salario_mensual=_dec(row[6]),
                    salario_semanal=_dec(row[7]),
                    eps=_s(row[9]) or None,
                    pension=_s(row[10]) or None,
                    arl=_s(row[11]) or None,
                    caja_compensacion=_s(row[12]) or None,
                ),
            )
            created += 1
        self.stdout.write(f"  Empleados: {created} procesados")

    # ── lotes ─────────────────────────────────────────────────────────────────

    def _seed_lotes(self, wb):
        ws = wb["Lotes y arboles"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        created = 0
        for row in rows:
            nombre = _s(row[0]) if row[0] else ""
            if not nombre or nombre not in REAL_LOTE_NAMES:
                continue
            # Use most recent tree count: col[5] = Palos agosto 17 2023
            num_arboles = int(row[5]) if row[5] and isinstance(row[5], (int, float)) and row[5] > 0 else 0
            variedad = _s(row[10]) if len(row) > 10 and row[10] else None
            ano_siembra = _s(row[11]) if len(row) > 11 and row[11] else None
            prox_renovacion = _s(row[12]) if len(row) > 12 and row[12] else None
            gramos = _dec(row[13]) if len(row) > 13 else None
            bultos_prod = _dec(row[14]) if len(row) > 14 else None
            bultos_urea = _dec(row[15]) if len(row) > 15 else None
            bultos_dap = _dec(row[16]) if len(row) > 16 else None
            Lote.objects.get_or_create(
                nombre=nombre,
                defaults=dict(
                    variedad=variedad,
                    año_siembra=ano_siembra,
                    proxima_renovacion=prox_renovacion,
                    num_arboles=num_arboles,
                    gramos_abono_palo=gramos,
                    bultos_produccion=bultos_prod,
                    bultos_urea=bultos_urea,
                    bultos_dap=bultos_dap,
                ),
            )
            created += 1
        self.stdout.write(f"  Lotes: {created} procesados")

    # ── observaciones ─────────────────────────────────────────────────────────

    def _seed_observaciones(self, wb):
        ws = wb["Observaciones"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            texto = _s(row[1]) if len(row) > 1 else ""
            if not fecha or not texto:
                continue
            objs.append(Observacion(fecha=fecha, observacion=texto))
        Observacion.objects.bulk_create(objs, ignore_conflicts=True)
        self.stdout.write(f"  Observaciones: {len(objs)}")

    # ── egresos ───────────────────────────────────────────────────────────────

    def _seed_egresos(self, wb):
        ws = wb["Egresos 2025"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        _prov_cache: dict[str, Proveedor | None] = {}
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            nombre = _s(row[1]) if row[1] else "Sin nombre"
            valor = _dec(row[6])
            if valor is None:
                continue
            cuenta_nombre = _s(row[7]) if row[7] else "Efectivo"
            cuenta = _get_cuenta(cuenta_nombre)
            if cuenta is None:
                cuenta = _get_cuenta("Efectivo")

            cat_raw = _s(row[3]) if row[3] else "varios"
            categoria = _map_categoria(cat_raw)

            prov_nombre = _s(row[8]) if len(row) > 8 and row[8] else None
            proveedor = None
            if prov_nombre:
                if prov_nombre not in _prov_cache:
                    _prov_cache[prov_nombre] = Proveedor.objects.filter(nombre=prov_nombre).first()
                proveedor = _prov_cache[prov_nombre]

            estado_raw = _s(row[13]).lower() if len(row) > 13 and row[13] else "pagada"
            estado = "pagada" if estado_raw in ("", "pagada") else ("pendiente" if "pendiente" in estado_raw else "pagada")

            objs.append(Egreso(
                fecha=fecha,
                nombre=nombre,
                descripcion=_s(row[2]) or None,
                cantidad=_dec(row[4]),
                unidad=_s(row[5]) or None,
                valor=valor,
                cuenta=cuenta,
                categoria=categoria,
                proveedor=proveedor,
                nit_proveedor_destino=_s(row[9]) or None if len(row) > 9 else None,
                facturado_a=_s(row[10]) or None if len(row) > 10 else None,
                abono_deuda=_dec(row[11], Decimal(0)) if len(row) > 11 else Decimal(0),
                restante=_dec(row[12], Decimal(0)) if len(row) > 12 else Decimal(0),
                estado=estado,
            ))

        Egreso.objects.bulk_create(objs, batch_size=500, ignore_conflicts=True)
        self.stdout.write(f"  Egresos: {len(objs)}")

    # ── ingresos ──────────────────────────────────────────────────────────────

    def _seed_ingresos(self, wb):
        ws = wb["Ingresos"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            descripcion = _s(row[1]) if row[1] else "Sin descripción"
            valor = _dec(row[2])
            if valor is None:
                continue
            cuenta_nombre = _s(row[3]) if row[3] else "Efectivo"
            cuenta = _get_cuenta(cuenta_nombre)
            objs.append(Ingreso(
                fecha=fecha,
                descripcion=descripcion,
                valor=valor,
                cuenta_destino=cuenta,
                origen=_s(row[4]) or None if len(row) > 4 else None,
                observaciones=_s(row[5]) or None if len(row) > 5 else None,
            ))

        Ingreso.objects.bulk_create(objs, batch_size=500, ignore_conflicts=True)
        self.stdout.write(f"  Ingresos: {len(objs)}")

    # ── transacciones ─────────────────────────────────────────────────────────

    def _seed_transacciones(self, wb):
        ws = wb["Transacciones"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            origen_nombre = _s(row[1]) if row[1] else None
            destino_nombre = _s(row[2]) if row[2] else None
            valor = _dec(row[3])
            if valor is None or not destino_nombre:
                continue

            # Set saldo_inicial for opening balance entries instead of creating transaccion
            obs = _s(row[4]) if len(row) > 4 else ""
            if origen_nombre is None and "saldo inicial" in obs.lower():
                cuenta = _get_cuenta(destino_nombre)
                if cuenta:
                    cuenta.saldo_inicial = valor
                    cuenta.save(update_fields=["saldo_inicial"])
                continue

            cuenta_origen = _get_cuenta(origen_nombre) if origen_nombre else None
            cuenta_destino = _get_cuenta(destino_nombre)
            objs.append(Transaccion(
                fecha=fecha,
                cuenta_origen=cuenta_origen,
                cuenta_destino=cuenta_destino,
                valor=valor,
                observaciones=obs or None,
            ))

        Transaccion.objects.bulk_create(objs, batch_size=300, ignore_conflicts=True)
        self.stdout.write(f"  Transacciones: {len(objs)}")

    # ── préstamos empleados ───────────────────────────────────────────────────

    def _seed_prestamos(self, wb):
        ws = wb["Préstamos Empleados"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        prestamos_created = 0
        abonos_created = 0

        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            valor = _dec(row[1])
            if valor is None:
                continue
            trabajador = _s(row[2]) if row[2] else None
            if not trabajador:
                continue

            empleado, _ = Empleado.objects.get_or_create(
                nombre_completo=trabajador,
                defaults={"activo": False},
            )

            # Parse abono pairs: (valor, fecha) at indices (5,6),(7,8),(9,10),(11,12),(13,14)
            abonos_data = []
            for i in range(5, 15, 2):
                if i + 1 >= len(row):
                    break
                abono_val = _dec(row[i])
                if abono_val is None:
                    continue
                abono_fecha = _date(row[i + 1])
                if abono_fecha is None:
                    abono_fecha = _date(row[4]) if row[4] else fecha
                abonos_data.append((abono_val, abono_fecha))

            total_abonado = sum(a[0] for a in abonos_data)
            saldo = valor - total_abonado

            prestamo = PrestamoEmpleado.objects.create(
                fecha=fecha,
                valor=valor,
                empleado=empleado,
                concepto=_s(row[3]) or "Sin concepto",
                saldo=saldo,
            )
            prestamos_created += 1

            for abono_val, abono_fecha in abonos_data:
                AbonoPrestamo.objects.create(
                    prestamo=prestamo,
                    fecha=abono_fecha,
                    valor=abono_val,
                )
                abonos_created += 1

        self.stdout.write(f"  Préstamos: {prestamos_created} | Abonos: {abonos_created}")

    # ── ventas café ───────────────────────────────────────────────────────────

    def _seed_ventas_cafe(self, wb):
        ws = wb["Ventas Café"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            kilos = _dec(row[1])
            valor_total = _dec(row[8])
            valor_neto = _dec(row[11])
            if kilos is None or valor_total is None or valor_neto is None:
                continue
            cargas = _dec(row[2], Decimal(0))
            cuenta_nombre = _s(row[12]) if len(row) > 12 and row[12] else "Efectivo"
            cuenta = _get_cuenta(cuenta_nombre)
            objs.append(VentaCafe(
                fecha=fecha,
                kilos=kilos,
                cargas=cargas,
                tipo_cafe=_map_tipo_cafe(_s(row[3])),
                factor=_dec(row[4]),
                precio_kilo=_dec(row[5], Decimal(0)),
                precio_carga=_dec(row[6]),
                comprador=_s(row[7]) or "Sin comprador",
                valor_total=valor_total,
                deduccion=_dec(row[9], Decimal(0)),
                retefuente=_dec(row[10], Decimal(0)),
                valor_neto=valor_neto,
                cuenta_destino=cuenta,
                facturado_a=_s(row[13]) or None if len(row) > 13 else None,
                conversion_cereza_seco=_dec(row[14]) if len(row) > 14 else None,
                beneficio=_s(row[15]) or None if len(row) > 15 else None,
                transportador=_s(row[16]) or None if len(row) > 16 else None,
                valor_transporte=_dec(row[17]) if len(row) > 17 else None,
            ))

        VentaCafe.objects.bulk_create(objs, batch_size=300, ignore_conflicts=True)
        self.stdout.write(f"  Ventas Café: {len(objs)}")

    # ── ventas banano ─────────────────────────────────────────────────────────

    def _seed_ventas_banano(self, wb):
        ws = wb["Ventas Banano"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            kilos = _dec(row[2])
            valor_total = _dec(row[4])
            if kilos is None or valor_total is None:
                continue
            tipo = _map_tipo_platano(_s(row[1]))
            precio_kilo = _dec(row[3], Decimal(0))
            cuenta_nombre = _s(row[5]) if len(row) > 5 and row[5] else "Efectivo"
            cuenta = _get_cuenta(cuenta_nombre)
            objs.append(VentaBanano(
                fecha=fecha,
                tipo_platano=tipo,
                kilos=kilos,
                precio_kilo=precio_kilo,
                valor_total=valor_total,
                cuenta_destino=cuenta,
                facturado_a=_s(row[6]) or None if len(row) > 6 else None,
                observaciones=_s(row[7]) or None if len(row) > 7 else None,
            ))

        VentaBanano.objects.bulk_create(objs, batch_size=300, ignore_conflicts=True)
        self.stdout.write(f"  Ventas Banano: {len(objs)}")

    # ── floraciones ───────────────────────────────────────────────────────────

    def _seed_floraciones(self, wb):
        ws = wb["Floraciones"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        _lote_cache: dict[str, Lote | None] = {}
        objs = []
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            calidad_raw = _s(row[2]) if len(row) > 2 and row[2] else "buena"
            if not fecha:
                continue
            lote_nombre = _s(row[1]) if len(row) > 1 and row[1] else None
            lote = None
            if lote_nombre:
                if lote_nombre not in _lote_cache:
                    _lote_cache[lote_nombre] = Lote.objects.filter(nombre=lote_nombre).first()
                lote = _lote_cache[lote_nombre]
            abonada = bool(row[3]) if len(row) > 3 and row[3] is not None else None
            broca = bool(row[4]) if len(row) > 4 and row[4] is not None else None
            roya = bool(row[5]) if len(row) > 5 and row[5] is not None else None
            objs.append(Floracion(
                fecha=fecha,
                lote=lote,
                calidad=_map_calidad(calidad_raw),
                abonada_ideal=abonada,
                broca_ideal=broca,
                roya_ideal=roya,
            ))

        Floracion.objects.bulk_create(objs, ignore_conflicts=True)
        self.stdout.write(f"  Floraciones: {len(objs)}")

    # ── mezclas abono ─────────────────────────────────────────────────────────

    def _seed_mezclas_abono(self, wb):
        ws = wb["Mezcla Abonos"]
        rows = list(ws.iter_rows(values_only=True))[1:]
        _lote_cache: dict[str, Lote | None] = {}
        mezclas = 0
        ferts = 0
        for row in rows:
            fecha = _date(row[0]) if row[0] else None
            if not fecha:
                continue
            formula = _s(row[1]) if row[1] else "Sin fórmula"
            lote_nombre = _s(row[2]) if len(row) > 2 and row[2] else None
            lote = None
            if lote_nombre:
                if lote_nombre not in _lote_cache:
                    _lote_cache[lote_nombre] = Lote.objects.filter(nombre=lote_nombre).first()
                lote = _lote_cache[lote_nombre]
            num_arboles = int(row[3]) if len(row) > 3 and row[3] and isinstance(row[3], (int, float)) else None
            gramos = _dec(row[4]) if len(row) > 4 else None
            costo_total = _dec(row[20]) if len(row) > 20 else None

            mezcla = MezclaAbono.objects.create(
                fecha=fecha,
                formula=formula,
                lote=lote,
                num_arboles=num_arboles,
                gramos_por_arbol=gramos,
                costo_total=costo_total,
            )
            mezclas += 1

            # Fertilizante pairs at (5,6,7), (8,9,10), (11,12,13), (14,15,16), (17,18,19)
            for base in range(5, 20, 3):
                if base + 1 >= len(row):
                    break
                nombre_fert = _s(row[base]) if row[base] else None
                num_bultos = _dec(row[base + 1]) if base + 1 < len(row) else None
                precio = _dec(row[base + 2]) if base + 2 < len(row) else None
                if not nombre_fert or num_bultos is None:
                    continue
                MezclaAbonoFertilizante.objects.create(
                    mezcla=mezcla,
                    fertilizante=nombre_fert,
                    num_bultos=num_bultos,
                    precio_bulto=precio,
                )
                ferts += 1

        self.stdout.write(f"  Mezclas Abono: {mezclas} | Fertilizantes: {ferts}")
